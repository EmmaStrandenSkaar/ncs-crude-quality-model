"""
Script 53 — Hent etterspørsels-data per oljeprodukt fra EIA.

KILDER:
  1. EIA STEO (Short-Term Energy Outlook) Excel — månedlige globale + US-data
     med 2-års forward forecasts. Tab 4a har US per-produkt konsum.
     URL: https://www.eia.gov/outlooks/steo/xls/STEO_m.xlsx

  2. EIA Weekly Product Supplied — ukentlige historiske data fra 1990-tallet
     Gasoline, distillate, jet fuel, residual fuel oil
     URL-mønster: https://www.eia.gov/dnav/pet/hist_xls/W{KODE}.xls

OUTPUT:
  data/raw/eia_demand/steo_full.xlsx        (cached, oppdater hver mnd)
  data/processed/53_product_demand.csv      (kombinert monthly demand-tabell)

KOLONNER I OUTPUT-CSV:
  date, source, motor_gasoline_kbpd, jet_fuel_kbpd, distillate_kbpd,
  residual_fuel_kbpd, total_petroleum_kbpd, is_forecast
"""

from pathlib import Path
import datetime as dt
import requests
import pandas as pd
import openpyxl

PROJECT_ROOT = Path(__file__).parent.parent
RAW_DIR      = PROJECT_ROOT / "data" / "raw" / "eia_demand"
PROC_DIR     = PROJECT_ROOT / "data" / "processed"
RAW_DIR.mkdir(parents=True, exist_ok=True)

# ── EIA STEO Excel ───────────────────────────────────────────────────────────
STEO_URL    = "https://www.eia.gov/outlooks/steo/xls/STEO_m.xlsx"
STEO_PATH   = RAW_DIR / "steo_full.xlsx"

# STEO tab 4a — rader (basert på inspisering)
STEO_4A_ROWS = {
    "motor_gasoline_mbpd":   46,
    "jet_fuel_mbpd":         47,
    "distillate_fuel_mbpd":  48,
    "residual_fuel_mbpd":    49,
    "total_petroleum_mbpd":  43,
}

# Inventory rows
STEO_4A_INVENTORY_ROWS = {
    "motor_gasoline_stocks_mbbl":  60,
    "jet_fuel_stocks_mbbl":        61,
    "distillate_stocks_mbbl":      62,
    "residual_stocks_mbbl":        63,
}

HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36"),
}


def download_steo(force: bool = False) -> Path:
    """Hent STEO Excel hvis ikke cached (eller hvis force=True)."""
    if STEO_PATH.exists() and not force:
        age_days = (dt.datetime.now() -
                    dt.datetime.fromtimestamp(STEO_PATH.stat().st_mtime)).days
        if age_days < 7:
            print(f"  · STEO cached ({age_days} dager gammel) — bruker eksisterende")
            return STEO_PATH

    print(f"  → Laster ned STEO Excel ({STEO_URL})...")
    r = requests.get(STEO_URL, headers=HEADERS, timeout=60)
    r.raise_for_status()
    STEO_PATH.write_bytes(r.content)
    print(f"  ✓ Lagret: {STEO_PATH.name} ({len(r.content)/1e6:.1f} MB)")
    return STEO_PATH


def parse_steo_4a(steo_path: Path) -> pd.DataFrame:
    """
    Parse STEO tab 4a til lange-format DataFrame.

    Returnerer: én rad per måned med kolonner for hver produkt-demand
    + inventory + et flagg `is_forecast` som markerer historiske vs prognose.
    """
    wb = openpyxl.load_workbook(steo_path, data_only=True)
    ws = wb["4atab"]

    # ── Bygg dato-vektor fra header (rad 3 = år, rad 4 = mnd) ───────────────
    year_row  = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    month_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    # Hver kolonne har år + måned. År gjentas ikke per måned — må forward-fill
    months = {"Jan":1, "Feb":2, "Mar":3, "Apr":4, "May":5, "Jun":6,
              "Jul":7, "Aug":8, "Sep":9, "Oct":10, "Nov":11, "Dec":12}
    dates = []
    current_year = None
    for col, (yr, mo) in enumerate(zip(year_row, month_row)):
        if isinstance(yr, (int, float)):
            current_year = int(yr)
        m_str = str(mo).strip() if mo else ""
        if m_str in months and current_year is not None:
            dates.append((col, dt.date(current_year, months[m_str], 1)))

    # ── Hent ut data per produkt ─────────────────────────────────────────────
    data_rows = []
    for label, row_n in {**STEO_4A_ROWS, **STEO_4A_INVENTORY_ROWS}.items():
        row = list(ws.iter_rows(min_row=row_n, max_row=row_n, values_only=True))[0]
        for col, date in dates:
            if col < len(row):
                val = row[col]
                if isinstance(val, (int, float)):
                    data_rows.append({"date": date, "variable": label, "value": float(val)})

    df = pd.DataFrame(data_rows)
    df = df.pivot(index="date", columns="variable", values="value").reset_index()
    df["date"] = pd.to_datetime(df["date"])

    # ── Bestem cutoff for forecast vs historisk ─────────────────────────────
    # STEO publiseres månedlig; "Forecast date" i celle A3 sier når den ble laget.
    forecast_date_cell = ws["A3"].value
    print(f"  STEO forecast date: {forecast_date_cell}")
    # Cutoff = nåværende måned. Alt etter = forecast
    today = dt.date.today()
    cutoff = dt.date(today.year, today.month, 1)
    df["is_forecast"] = df["date"].dt.date >= cutoff

    return df


def fetch_weekly_product_supplied() -> pd.DataFrame:
    """
    Ukentlig US product supplied (1991+) fra EIA dnav.
    Vi aggregerer til månedlig snitt for å matche STEO.
    """
    series = {
        "motor_gasoline_kbpd":  "WGFUPUS2",   # Finished motor gasoline
        "distillate_kbpd":       "WDIUPUS2",   # Distillate fuel oil
        "jet_fuel_kbpd":         "WKJUPUS2",   # Kerosene-type jet fuel
        "residual_fuel_kbpd":   "WRFUPUS2",   # Residual fuel oil
    }
    all_dfs = []
    for col, code in series.items():
        url = f"https://www.eia.gov/dnav/pet/hist_xls/{code}w.xls"
        print(f"  → {col:<30} ({code})")
        try:
            r = requests.get(url, headers=HEADERS, timeout=60)
            r.raise_for_status()
            local = RAW_DIR / f"weekly_{code}.xls"
            local.write_bytes(r.content)

            # Parse Excel — Data 1-arket har de ukentlige verdiene
            wb = openpyxl.load_workbook(local, data_only=True)
            sheet = wb["Data 1"]
            rows = []
            header_seen = False
            for row in sheet.iter_rows(values_only=True):
                if not header_seen:
                    if row[0] and "Date" in str(row[0]):
                        header_seen = True
                    continue
                if isinstance(row[0], (dt.datetime, dt.date)) and isinstance(row[1], (int, float)):
                    rows.append({"date": row[0], "value": float(row[1])})
            df = pd.DataFrame(rows)
            df = df.rename(columns={"value": col})
            all_dfs.append(df.set_index("date")[[col]])
            print(f"     {len(df)} ukentlige obs")
        except Exception as e:
            print(f"     ⚠ Feilet: {e}")

    if not all_dfs:
        return pd.DataFrame()

    weekly = pd.concat(all_dfs, axis=1, join="outer").sort_index()
    # Aggreger til månedlig snitt
    monthly = weekly.resample("MS").mean().reset_index()
    monthly = monthly.rename(columns={"date": "date"})
    return monthly


def main():
    print("=" * 70)
    print("  SCRIPT 53: Hent EIA produkt-demand data")
    print("=" * 70)

    # ── STEP 1: STEO Excel (historisk + forecast) ────────────────────────────
    print("\n[1] STEO Excel (US per-product consumption with 2-year forecast)")
    steo_path = download_steo()
    steo_df = parse_steo_4a(steo_path)
    print(f"  ✓ STEO 4a parset: {len(steo_df)} mnd, "
          f"{steo_df['date'].min().strftime('%Y-%m')} → {steo_df['date'].max().strftime('%Y-%m')}")
    print(f"  Hist: {(~steo_df['is_forecast']).sum()} mnd, "
          f"Forecast: {steo_df['is_forecast'].sum()} mnd")

    # Konverter mbpd → kbpd for konsistens
    for col in steo_df.columns:
        if col.endswith("_mbpd"):
            steo_df[col.replace("_mbpd", "_kbpd")] = steo_df[col] * 1000
            steo_df = steo_df.drop(col, axis=1)

    # ── STEP 2: Weekly product supplied (lengre historikk) ──────────────────
    print("\n[2] EIA Weekly Product Supplied (US, 1990+, månedlig aggregert)")
    weekly_df = fetch_weekly_product_supplied()
    if not weekly_df.empty:
        print(f"  ✓ Weekly data: {len(weekly_df)} mnd, "
              f"{weekly_df['date'].min().strftime('%Y-%m')} → "
              f"{weekly_df['date'].max().strftime('%Y-%m')}")

    # ── STEP 3: Kombiner — historisk weekly (hvis tilgjengelig) + STEO ─────
    print("\n[3] Kombinerer datakilder...")

    # STEO: alltid hovedkilden — har både historikk og forecast
    steo_df["source"] = steo_df["is_forecast"].map({
        True:  "EIA STEO (forecast)",
        False: "EIA STEO (historical)"
    })

    if not weekly_df.empty:
        # Weekly har høyere presisjon for historiske perioder → bruk det
        weekly_renamed = weekly_df.rename(columns={
            "distillate_kbpd": "distillate_fuel_kbpd",
        })
        weekly_renamed["source"]      = "EIA Weekly (actual)"
        weekly_renamed["is_forecast"] = False

        # Weekly har prioritet for historikk → konkat + drop duplikater
        combined = pd.concat([weekly_renamed, steo_df], ignore_index=True)
        combined = combined.sort_values(["date", "source"])
        combined = combined.drop_duplicates(subset="date", keep="first")
    else:
        # Ingen weekly → bruk STEO alene (har 52 hist + 20 forecast)
        print("  · Weekly data utilgjengelig — bruker STEO alene")
        combined = steo_df.copy()

    # ── STEP 4: Lagre ────────────────────────────────────────────────────────
    out_path = PROC_DIR / "53_product_demand.csv"
    combined.to_csv(out_path, index=False)
    print(f"\n  ✓ Lagret: {out_path.name} ({len(combined)} mnd-rader)")
    print(f"  Hist (faktisk):    {(combined['source']=='EIA Weekly (actual)').sum()} mnd")
    print(f"  Forecast (STEO):   {(combined['source']=='EIA STEO (forecast)').sum()} mnd")

    # ── Oppsummering: gjennomsnitt per produkt over 2025 ────────────────────
    print(f"\n  2025 gjennomsnitt (US, kbpd):")
    recent = combined[(combined["date"] >= "2025-01-01") &
                       (combined["date"] <= "2025-12-31")]
    for col in ["motor_gasoline_kbpd", "jet_fuel_kbpd",
                 "distillate_fuel_kbpd", "residual_fuel_kbpd"]:
        if col in recent.columns:
            m = recent[col].mean()
            print(f"    {col:<28} {m:>8,.0f} kbpd")

    # Forecast trajectory
    print(f"\n  Forecast-trajectory (gjennomsnitt per år, kbpd):")
    print(f"  {'År':<6} {'Gasoline':>10} {'Jet':>8} {'Distillate':>11} {'Residual':>9}")
    for yr in [2024, 2025, 2026, 2027]:
        sub = combined[combined["date"].dt.year == yr]
        if len(sub) > 0:
            print(f"  {yr:<6} "
                  f"{sub['motor_gasoline_kbpd'].mean():>10,.0f} "
                  f"{sub['jet_fuel_kbpd'].mean():>8,.0f} "
                  f"{sub['distillate_fuel_kbpd'].mean():>11,.0f} "
                  f"{sub['residual_fuel_kbpd'].mean():>9,.0f}")


if __name__ == "__main__":
    main()
